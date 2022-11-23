import openpyxl
# wb = openpyxl.Workbook('E:\PYTHON临时\评论excel\Book1.xlsx')
# wb1 = openpyxl.load_workbook('ff.xls')

# print(wb)
# print(type(wb))
# print(wb1)
# print(type(wb1))


# from openpyxl import load_workbook
# book = openpyxl.Workbook()
# uid = "5513020355"
# # fileName 这里是指文件路径
# fileName = "ff.xlsx"
# # 以只读模式打开工作簿
# wb = load_workbook(filename = fileName,read_only = True)
# # sheetName 就是 sheet页的名称
# sheetName = book.create_sheet(uid)
# # 通过 工作表名 获取 工作表
# ws = wb[sheetName]
# # 按行读取 工作表的内容
# for row in ws.rows:
#     for cell in row:
#         # 输出 单元格中的数据
#         print(cell.value)


def make_one_sheet(uid,book):
    global data,sheet
    data = openpyxl.load_workbook('E:\PYTHON临时\评论excel\{}.xlsx'.format(uid)) #读取格式转换后的文件三
    sheet = book.create_sheet(uid)
    col = ['评论序号','评论内容','评论时间']
    for i in range(1,3):
        sheet.cell(row=1,column=i).value=col[i]



import os

import win32com.client as win32

file_path = r"E:\PYTHON临时\评论excel"


def to_xlsx(file):
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    wb = excel.Workbooks.Open(file)
    wb.SaveAs(file + 'x', FileFormat=51)  # FileFormat = 51 is for .xlsx extension
    wb.Close()  # FileFormat = 56 is for .xls extension
    excel.Application.Quit()


if __name__ == '__main__':
    os_listdir = os.listdir(file_path)
    for i in os_listdir:
        file = os.path.join(file_path, i)
        to_xlsx(file)