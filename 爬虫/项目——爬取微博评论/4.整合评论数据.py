# import jieba
# import re
import openpyxl
# from openpyxl import load_workbook
import pandas as pd


data= pd.read_excel("E:\\PYTHON临时\\评论excel\\5513020355.xlsx")
# data = r"E:\PYTHON临时\评论excel\5513020355.xlsx"
newbook=None
sheet=None
redundancy=["<a href=","<img alt="]

def make_one_sheet(uid,book):
    global data,sheet
    data = openpyxl.load_workbook('E:\\PYTHON临时\\评论excel\\5513020355.xlsx') #读取格式转换后的文件三  .format(uid)
    sheet = book.create_sheet(uid)
    col = ['评论序号','评论内容','评论时间']
    for i in range(1,3):
        sheet.cell(row=1,column=i).value=col[i]

    # return data,sheet

def fill_one_sheet():
    global data,sheet
    count = len(data.sheetnames)
    k = 2
    for j in range(count):
        table = data.get_sheet_by_name('微博{}'.format(j+1))
        rows = table.max_row
        print('微博{}一共有{}行数据，开始清洗数据...'.format(j+1,rows))
        for i in range(2,rows):
            Comment = str(table.cell(row=i,column=2).value)
            Time = str(table.cell(row=i,column=3).value)
            if len(Time)>19:#时间规范化
                Time=Time[:19]
            if ("<a href=" in Comment) or ("<img alt=" in Comment):#去除因引用、回复或插入图片、表情而来的冗余文本
                head1, sep1, tail1 =Comment.partition("<")
                head2, sep2, tail2 =tail1.partition(">")
                Comment=head1+tail2
                Comment=Comment.replace("评论配图","")
            if ("@" in Comment) :
                head1, sep1, tail1 =Comment.partition("@")
                head2, sep2, tail2 =tail1.partition(":")
                Comment=head1+tail2
            if Comment != "":
                sheet.cell(row=k,column=1).value=k-1
                sheet.cell(row=k,column=2).value=Comment
                sheet.cell(row=k,column=3).value=Time
                k=k+1

def save_excel(book,uid):
    savepath = 'E:\PYTHON临时\评论excel\{}整合.xlsx'.format(uid)
    book.save(savepath)

def main():
    uid="5513020355"
    # data = r"E:\PYTHON临时\评论excel\5513020355.xlsx"
    newbook=openpyxl.Workbook()
    # make_one_sheet(uid,newbook)
    fill_one_sheet()
    save_excel(newbook,uid)

main()
